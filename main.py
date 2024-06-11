import os
import re
from PyPDF2 import PdfReader
from crewai import Agent, Task, Process, Crew
from crewai_tools import tool
import openpyxl
from openpyxl import load_workbook

file_path = "resume_data.xlsx"
os.environ["OPENAI_MODEL_NAME"] = 'gpt-4'  
os.environ["OPENAI_API_KEY"] = 'Your API Key'

@tool
def write_to_excel_tool(file_path: str, data: list) -> str:
    """
    Writes the provided data to an Excel file at the specified file path.
    If the file exists, it appends the data to it.
    If the file does not exist, it creates a new file and writes the data to it.

    Args:
        file_path (str): The path where the Excel file will be saved.
        data (list): A list of dictionaries containing the data to be written to the Excel file.

    Returns:
        str: A message indicating the success of the operation.
    """
    if not data:
        return "No data provided to write to the Excel file."

    headers = list(data[0].keys())

    try:
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            existing_headers = [cell.value for cell in worksheet[1]]

            if existing_headers != headers:
                return "Headers do not match. Please ensure the data format is consistent."

            start_row = worksheet.max_row + 1
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for col_num, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col_num, value=header)
            start_row = 2

        for row_num, row_data in enumerate(data, start_row):
            for col_num, (key, cell_data) in enumerate(row_data.items(), 1):
                worksheet.cell(row=row_num, column=col_num, value=cell_data)

        workbook.save(file_path)
        return f"Data successfully written to {file_path}"

    except PermissionError:
        return f"Permission denied: unable to write to {file_path}"
    except Exception as e:
        return f"An error occurred: {e}"

def cache_func(args, result):
    return True  
write_to_excel_tool.cache_function = cache_func

def extract_text_from_pdf(pdf_path):
    pdf_reader = PdfReader(pdf_path)
    text = ""
    for page in pdf_reader.pages:
        text_1 = page.extract_text()
        if text_1:
            text_1 = re.sub(r"(\w+)-\n(\w+)", r"\1\2", text_1)
            text_1 = re.sub(r"(?<!\n\s)\n(?!\s\n)", " ", text_1.strip())
            text_1 = re.sub(r"\n\s*\n", "\n\n", text_1)
            text += text_1
    return text
def process_resume(resume_text):
    category_agent = Agent(
        role="categorizer",
        goal=f"You should categorize the {resume_text} into two different categories 'Fresher' and 'Experience' by analysing the resume properly and categorize it accurately by using the keywords like having a company name and years of experience. If nothing that type of information is present, then consider it as a fresher. Search for keywords or phrases indicating past employment, such as 'company name', 'employment history', 'work experience', etc. If any company names are found, categorize the resume as 'Experienced'. If company names are found with specific dates or durations, categorize the resume as 'Experienced' and don't get confused between the qualifications like the education with the company. Both are different. Make sure you properly understand and then process by their name. Use a predefined list of keywords or employ Natural Language Processing (NLP) techniques for more comprehensive extraction and make sure you present the person's name of the resume perfectly and also their details like phone number and gmail and gender with the name of the person. Analyze properly. This data is available for sure in the file, so mention these things while categorizing.",
        backstory="Your are an AI assistant whose only job is to categorize the resume into two different categories 'Fresher' and 'Experience' and nothing else.",
        verbose=True,
        allow_delegation=False
    )

    entity_agent = Agent(
        role="entity finder",
        goal="You should find out the keywords that can tell the person's capabilities like 'skills', 'programming languages', 'projects' and the summary of that person and consider other things which may show their commitment level and other things. This is for 'Fresher' and for experience the same things are needed but there are some extra things like experience and the projects they worked on and also the company which they worked and the years they worked. You have to only bring those kinds of attributes. 1. Keyword Extraction: Identify keywords related to skills, programming languages, projects, summary, commitment level, and experience (if applicable). Use a predefined list of keywords or employ Natural Language Processing (NLP) techniques for more comprehensive extraction. 2. Fresher vs. Experienced Classification: Company and Experience Indicators: Search for keywords like 'company name', 'employment history', 'work experience', etc. If company names are found with specific dates or durations, categorize as 'Experienced'. Default Fresher Classification: If no company names or experience indicators are found, categorize as 'Fresher'. 3. Attribute Extraction: Skills and Projects: Identify keywords and phrases related to skills and projects. Extract project details like title, description, and technologies used (if available). Summary: Extract the summary section of the resume (if present). (For Experienced only): Experience: Identify and extract details for each work experience, including company name, duration (dates or years), and job title. 4. Optimization: Language Independence: Include keywords in relevant languages for broader applicability. Customizable Keyword Lists: Allow customization of skill and experience keywords based on specific needs. Confidence Scores (Optional): Implement techniques to assign confidence scores to extracted attributes.",
        backstory="Your are an AI assistant whose only job is to find the keywords which help determine the person's capabilities.",
        verbose=True,
        allow_delegation=False
    )

    score_agent = Agent(
        role="Score shortlist",
        goal=f"You should assign scores to the keywords based on how much impact those keywords can have, assigning proper scores. The shortlisting minimum score is 75, and how much weight they hold. Define Weights: Freshers: Weight skills, projects, and summary heavily, reflecting their potential and academic achievements. Years of experience won't be considered. Experienced: Weight skills, project experience, and years of relevant experience more heavily. Company reputation can be a bonus factor. Skill Matching: Apply the same logic as before, awarding points for matching skills based on their weight for the specific category (Fresher or Experienced). Project Evaluation: Freshers: Focus on project relevance to the job and the demonstrated skills. Award points based on the complexity and impact of the project (e.g., personal vs. academic). Experienced: Consider project relevance, scale, impact, and technologies used. Assign higher points for projects directly related to the job requirements. Experience (Experienced only): Award points based on the total years of relevant experience. Summary Analysis (Optional): Apply the same logic as before, awarding a small bonus score for positive indicators of motivation and relevant goals. Aggregate Score: Freshers: Sum the points from skills, projects, and summary. Experienced: Sum the points from skills, projects, experience, and (optional) company reputation bonus. Normalization: Normalize the scores within a specific range (e.g., 0-100) for both Fresher and Experienced categories to allow for easier comparison within each group. Thresholds: Define separate minimum score thresholds for shortlisting Freshers and Experienced candidates based on their respective scoring systems. Use a predefined list of keywords or employ Natural Language Processing (NLP) techniques for more comprehensive evaluation and give scores accordingly with proper format. Don't just give scores different each time, keep fixed scores and give to them. Be strict while evaluating the scores because everyone can shortlist with the scores if you give randomly to everything, so be strict. Use good mathematical techniques and evaluation with high precision and proper classification.",
        backstory="Your are an AI assistant whose only job is to assign scores to it and shortlist them.",
        verbose=True,
        allow_delegation=False
    )

    data_agent = Agent(
        role="data extractor",
        goal=f"You are to classify the data {resume_text} into dictionary format from the output of their work. The things you need to classify into dictionary format are 'name','Gender', 'gmail', 'address', 'phone number or contact number', and 'category 'Experience or fresher'' and  you need to classify into dictionary format are 'skills', 'programming languages', 'Degree','Degree Stream', 'Year of pass out','10th percentage','12th percentage','Degree Percentage','Internship Experience','Job Experience','LinkedIn ID','GitHub ID','projects', 'summary', into the dictionary format and from the data, I need only 'Shortlisted as 'yes' or 'no''. If anywhere for the attribute values is not present, then place 'NA' term in it but check properly that the value is present there will the value for sure chech properly and analysis properly and search there will the values present for all the attributes chech with high presision if not then just place NA term in the attribute but check with high presision and do the job and wiith a well structured format and prefectly. Make sure to classify and format the data properly with high precision and proper data as the output. Use a predefined list of keywords or employ Natural Language Processing (NLP) techniques for more comprehensive extraction and evaluation of the data properly with perfection.",
        backstory="You are an AI assistant whose only job is to classify the given attributes data into dictionary format with perfection and high precision.",
        verbose=True,
        allow_delegation=False
    )

    category_task = Task(
        description="categorizing the resume as 'Fresher' and 'Experience'.",
        agent=category_agent,
        expected_output="classifying the resumes 'Fresher' and 'Experience' with their resume separately"
    )

    entity_task = Task(
        description="extracting the keywords which help determine the capabilities of that person",
        agent=entity_agent,
        expected_output="getting all the keywords like skills, projects, experience based on the person is 'Fresher' or 'Experience'"
    )

    score_task = Task(
        description="assigning scores to the keywords accordingly and shortlisting the resume based on the scores.",
        agent=score_agent,
        expected_output="assigning scores to the keywords and shortlisting"
    )

    data_task = Task(
        description="classifying the mentioned attributes into dictionary format.",
        agent=data_agent,
        expected_output="a proper dictionary formatted data of the requested attributes from the agents"
    )

    crew = Crew(
        agents=[category_agent, entity_agent, score_agent, data_agent],
        tasks=[category_task, entity_task, score_task, data_task],
        verbose=2,
        process=Process.sequential
    )
    
    output = crew.kickoff()
    return output

pdf_directory = "resumes"
pdf_files = [f for f in os.listdir(pdf_directory) if f.endswith(".pdf")]

output_results = []
for index, pdf_file in enumerate(pdf_files):
    pdf_path = os.path.join(pdf_directory, pdf_file)
    print(f"Processing file {index + 1}/{len(pdf_files)}: {pdf_file}")
    resume_text = extract_text_from_pdf(pdf_path)
    result = process_resume(resume_text)
    output_results.append(result)
    print(f"Result for {pdf_file}: {result}\n")

table_agent = Agent(
    role="Table creator",
    goal=f"You are to create a table and save it in an Excel format file named {file_path} using the data {output_results} the data contains list of dictionary formatted data you need to extract the data properly and use the data accordingly and prefectly with high presision and use it to create a prefect structured table and with the proper entry of the data and use the tool propvided as the sysntax and modify the code accordingly for the usage and make the file save in excel format prefectly. The table should include columns for  you need to classify are 'name','Gender', 'gmail', 'address', 'phone number or contact number', and 'category 'Experience or fresher' 'skills', 'programming languages', 'Degree','Degree Stream', 'Year of pass out','10th percentage','12th percentage','Degree Percentage','Internship Experience','Job Experience','LinkedIn ID','GitHub ID','projects', 'summary',  and from the data prefectly with high presion and use natural language methods for the process and make it proper with prefection. Each row should represent a resume. Ensure the data is accurate and formatted properly with high precision I need only 'Shortlisted as 'yes' or 'no''. If anywhere for the attribute values is not present, then place 'NA' term in it but check properly that the value is present there will the value for sure chech properly and analysis properly and search there will the values present for all the attributes chech with high presision if not then just place NA term in the attribute but check with high presision and do the job and wiith a well structured format and prefectly. Make sure to classify and format the data properly with high precision and proper data as the output. Use a predefined list of keywords or employ Natural Language Processing (NLP) techniques for more comprehensive extraction and evaluation of the data properly with perfection.",
    backstory="You are an AI assistant whose only job is to create a table with this data properly with high precision and perfection.",
    verbose=True,
    allow_delegation=False,
    tools=[write_to_excel_tool]
)

table_task = Task(
    description=f"Creating a table with this {output_results} data using write_to_excel_tool and saving it in an Excel format.",
    agent=table_agent,
    expected_output=f"Created a table with the data and saved it in an Excel format file named {file_path} with all the attributes mentioned."
)

crew = Crew(
    agents=[table_agent],
    tasks=[table_task],
    verbose=2,
    process=Process.sequential
)

output = crew.kickoff()
output
